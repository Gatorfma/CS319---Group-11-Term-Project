from django.contrib import messages  # ✔️ bu olacak

import pandas as pd
from django.shortcuts import redirect, render

from courses.models import Course, CourseOffering

from accounts.models import CustomUser, InstructorProfile, Student, TAProfile
from .forms import UploadExcelForm
import openpyxl
from django.http import HttpResponse

# Create your views here.
def import_data_view(request):
    def parse_semester(semester_code):
        year = int(str(semester_code)[:4])
        season_code = str(semester_code)[-1]
        season_map = {'1': 'Fall', '2': 'Spring', '3': 'Summer'}
        season = season_map.get(season_code, 'Fall')
        return f"{season} {year}"

    def safe_str(val):
        return str(val).strip() if pd.notna(val) else ''

    def safe_bool(val):
        return bool(val) if pd.notna(val) else False

    def safe_int(val, default=0):
        return int(val) if pd.notna(val) else default

    if request.method == 'POST':
        import_type = request.POST.get("import_type")
        file_key = f"{import_type}_file"
        uploaded_file = request.FILES.get(file_key)

        if not uploaded_file:
            messages.error(request, "No file selected.")
            return redirect('courses:import_data')

        try:
            df = pd.read_excel(uploaded_file)

            if import_type == "courses":
                df = df.rename(columns={
                    'department_code': 'department_code',
                    'course_no': 'course_code',
                    'course_name': 'title'
                })

                for _, row in df.iterrows():
                    department_code = safe_str(row['department_code']).upper()
                    course_code = safe_int(row['course_code'])
                    title = safe_str(row['title'])

                    Course.objects.get_or_create(
                        department_code=department_code,
                        course_code=course_code,
                        defaults={'title': title}
                    )

            elif import_type == "faculty":
                df = df.rename(columns={'department_code': 'department'})
                for _, row in df.iterrows():
                    department = safe_str(row['department']).upper()
                    staff_id = safe_str(row['staff_id'])
                    first_name = safe_str(row['first_name'])
                    last_name = safe_str(row['last_name'])
                    email = safe_str(row['email'])

                    user, created = CustomUser.objects.get_or_create(
                        employee_id=staff_id,
                        defaults={
                            'username': f"instructor{staff_id}",
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'role': CustomUser.Roles.INSTRUCTOR,
                            'department': department
                        }
                    )

                    if created:
                        user.set_password("defaultpass123")
                    else:
                        user.first_name = first_name
                        user.last_name = last_name
                        user.email = email
                        user.role = CustomUser.Roles.INSTRUCTOR
                        user.department = department

                    user.save()

                    if not hasattr(user, 'instructor_profile'):
                        InstructorProfile.objects.create(
                            user=user,
                            is_faculty=safe_bool(row.get('is_faculty', True)),
                            is_active=safe_bool(row.get('is_active', True))
                        )
                    else:
                        profile = user.instructor_profile
                        profile.is_faculty = safe_bool(row.get('is_faculty', True))
                        profile.is_active = safe_bool(row.get('is_active', True))
                        profile.save()

            elif import_type == "students":
                df = df.rename(columns={'department_code': 'department', 'class': 'student_class'})
                for _, row in df.iterrows():
                    student_id = safe_str(row['student_id'])
                    first_name = safe_str(row['first_name'])
                    last_name = safe_str(row['last_name'])

                    student, created = Student.objects.get_or_create(
                        student_id=student_id,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name
                        }
                    )

                    if not created:
                        if student.first_name != first_name or student.last_name != last_name:
                            student.first_name = first_name
                            student.last_name = last_name
                            student.save()

                    if safe_str(row.get('is_ta')) == '1':
                        department = safe_str(row['department']).upper()
                        email = safe_str(row['email'])
                        student_class = safe_int(row['student_class'])
                        ta_type = 'PHD' if student_class == 9 else 'GRAD'

                        user, created = CustomUser.objects.get_or_create(
                            username=student_id,
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'email': email,
                                'role': CustomUser.Roles.TA,
                                'department': department
                            }
                        )

                        if created:
                            user.set_password("defaultpass123")
                        else:
                            user.first_name = first_name
                            user.last_name = last_name
                            user.email = email
                            user.department = department
                            user.role = CustomUser.Roles.TA

                        user.employee_id = student_id
                        user.phone_number = safe_str(row.get('phone_no'))
                        user.save()

                        if not hasattr(user, 'ta_profile'):
                            TAProfile.objects.create(
                                user=user,
                                ta_type=ta_type,
                                is_active=safe_bool(row.get('is_active', True)),
                                proctor_type=safe_int(row.get('proctor_type'), 0)
                            )
                        else:
                            profile = user.ta_profile
                            profile.ta_type = ta_type
                            profile.is_active = safe_bool(row.get('is_active', True))
                            profile.proctor_type = safe_int(row.get('proctor_type'), 0)
                            profile.save()

            elif import_type == "course_offerings":
                df = df.rename(columns={'department_code': 'department'})
                for _, row in df.iterrows():
                    department = safe_str(row['department']).upper()
                    course = Course.objects.filter(department_code=department, course_code=safe_int(row['course_no'])).first()
                    if not course:
                        messages.warning(request, f"Course not found: {department}{row['course_no']}")
                        continue
                    semester = parse_semester(row['semester'])
                    section = safe_int(row['section_no'])
                    staff_id = safe_str(row['staff_id'])
                    user, _ = CustomUser.objects.get_or_create(
                        employee_id=staff_id,
                        defaults={
                            'username': f"instructor{staff_id}",
                            'first_name': 'Auto',
                            'last_name': 'Generated',
                            'role': CustomUser.Roles.INSTRUCTOR,
                            'department': department
                        }
                    )
                    user.set_password("defaultpass123")
                    user.save()

                    offering, _ = CourseOffering.objects.get_or_create(
                        course=course,
                        semester=semester,
                        section=section,
                        defaults={
                            'max_capacity': 100,
                            'enrolled_count': 0,
                        }
                    )
                    offering.instructors.add(user)
                    if not hasattr(user, 'instructor_profile'):
                        InstructorProfile.objects.create(user=user)

                    if hasattr(user, 'instructor_profile'):
                        user.instructor_profile.assigned_course_offerings.add(offering)

            elif import_type == "enrollments":
                df = df.rename(columns={'department_code': 'department'})
                for _, row in df.iterrows():
                    department = safe_str(row['department']).upper()
                    student_id = safe_str(row['student_id'])
                    semester = parse_semester(row['semester'])
                    section = safe_int(row['section_no'])
                    course = Course.objects.filter(department_code=department, course_code=safe_int(row['course_no'])).first()
                    if not course:
                        messages.warning(request, f"Course not found: {department}{row['course_no']}")
                        continue
                    offering = CourseOffering.objects.filter(course=course, semester=semester, section=section).first()
                    if not offering:
                        messages.warning(request, f"Offering not found for {course} - {semester} Sec {section}")
                        continue
                    student = Student.objects.filter(student_id=student_id).first()
                    if not student:
                        messages.warning(request, f"Student not found: {student_id}")
                        continue
                    offering.students.add(student)
                    student.enrolled_courses.add(offering)
                    offering.enrolled_count = offering.students.count()
                    offering.save()

                    try:
                        user = CustomUser.objects.get(username=student_id)
                        if hasattr(user, 'ta_profile'):
                            user.ta_profile.enrolled_course_offerings.add(offering)
                    except CustomUser.DoesNotExist:
                        pass

            messages.success(request, f"{import_type.replace('_', ' ').title()} imported successfully.")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f"Import failed: {e}")

        return redirect('courses:import_data')

    import_types = [
        ("courses", "Courses"),
        ("faculty", "Faculty"),
        ("students", "Students"),
        ("course_offerings", "Course Offerings"),
        ("enrollments", "Enrollments"),
    ]
    return render(request, 'courses/import_data.html', {"import_types": import_types})

def upload_student_excell(request):
    data = []
    classrooms = []
    assigned_classrooms = {}
    error = None

    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            classrooms = form.cleaned_data['classroom_selection']  # This is a queryset

            if not excel_file.name.endswith('.xlsx'):
                error = "Please upload a valid .xlsx file."
            else:
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    sheet = wb.active

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        name, surname, student_id = row
                        data.append({
                            'name': name,
                            'surname': surname,
                            'id': student_id
                        })

                    data.sort(key=lambda x: x['surname'])

                    assigned_classrooms = {cls: [] for cls in classrooms}
                    num_classrooms = len(classrooms)
                    students_per_classroom = len(data) // num_classrooms
                    remainder = len(data) % num_classrooms

                    start = 0
                    for i, classroom in enumerate(classrooms):
                        extra = 1 if i < remainder else 0  # Distribute remaining students fairly
                        end = start + students_per_classroom + extra
                        assigned_classrooms[classroom] = data[start:end]
                        start = end
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Student Assignments"

                    row_num = 1
                    for classroom, student_list in assigned_classrooms.items():
                        ws.cell(row=row_num, column=1, value=f"Classroom: {classroom}")
                        row_num += 1
                        ws.cell(row=row_num, column=1, value="Name")
                        ws.cell(row=row_num, column=2, value="Surname")
                        ws.cell(row=row_num, column=3, value="ID")
                        row_num += 1

                        for student in student_list:
                            ws.cell(row=row_num, column=1, value=student['name'])
                            ws.cell(row=row_num, column=2, value=student['surname'])
                            ws.cell(row=row_num, column=3, value=student['id'])
                            row_num += 1

                        row_num += 2  # space between classrooms

                    # Prepare response
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=assigned_students.xlsx'

                    # Save workbook to response
                    wb.save(response)
                    return response

                except Exception as e:
                    error = f"Failed to read Excel file: {str(e)}"
    else:
        form = UploadExcelForm()

    return render(request, 'courses/upload_student_excell.html', {
        'form': form,
        'data': data,
        'classrooms': classrooms,
        'assigned_classrooms': assigned_classrooms,
        'error': error})