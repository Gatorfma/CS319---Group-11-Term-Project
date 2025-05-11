from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DetailView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Prefetch
from django.utils import timezone
import csv
import io
from openpyxl import Workbook

from courses.models import Course

from .models import (
    TARequest, Semester, TAPreference, GraderPreference, 
    MustHaveTAPreference, AvoidTAPreference, AvoidGraderPreference
)
from .forms import TARequestForm, SemesterForm
from accounts.models import CustomUser, TAProfile, InstructorProfile


class InstructorRequiredMixin(UserPassesTestMixin):
    """Mixin to ensure only instructors can access the view"""
    
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_instructor()


class CoordinatorRequiredMixin(UserPassesTestMixin):
    """Mixin to ensure only TA coordinators can access the view"""
    
    def test_func(self):
        if not self.request.user.is_authenticated:
            return False
        if not self.request.user.is_instructor():
            return False
        try:
            return self.request.user.instructor_profile.is_ta_coordinator
        except:
            return False


class TARequestListView(LoginRequiredMixin, InstructorRequiredMixin, ListView):
    """List all TA requests created by the instructor"""
    model = TARequest
    template_name = 'ta_request_cs_dept/request_list.html'
    context_object_name = 'ta_request_cs_dept'
    
    def get_queryset(self):
        return TARequest.objects.filter(instructor=self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get active semester info
        try:
            active_semester = Semester.objects.filter(is_active=True).first()
            context['active_semester'] = active_semester
            context['is_request_period_active'] = active_semester and active_semester.is_request_period_active()
        except:
            context['is_request_period_active'] = False
            context['active_semester'] = None
        
        # Get courses without requests this semester
        if context['active_semester']:
            instructor_courses = Course.objects.all()
            existing_requests = TARequest.objects.filter(
                instructor=self.request.user,
                semester=active_semester
            ).values_list('course_id', flat=True)
            
            context['available_courses'] = instructor_courses.exclude(id__in=existing_requests)
        
        return context


class TARequestCreateView(LoginRequiredMixin, InstructorRequiredMixin, CreateView):
    """Create a new TA request"""
    model = TARequest
    form_class = TARequestForm
    template_name = 'ta_request_cs_dept/request_form.html'
    success_url = reverse_lazy('ta_request_cs_dept:list')
    
    def dispatch(self, request, *args, **kwargs):
        # Check if request period is active
        try:
            active_semester = Semester.objects.get(is_active=True)
            if not active_semester.is_request_period_active():
                messages.error(request, f"TA request submission is closed. The request period ended on {active_semester.request_end_date}.")
                return redirect('ta_request_cs_dept:list')
        except Semester.DoesNotExist:
            messages.error(request, "No active semester found. Please contact the administrator.")
            return redirect('ta_request_cs_dept:list')
            
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instructor'] = self.request.user
        kwargs['semester'] = Semester.objects.filter(is_active=True).first()
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tas'] = TAProfile.objects.filter(is_active=True)
        context['semester'] = Semester.objects.filter(is_active=True).first()
        return context
    
    def form_valid(self, form):
        form.instance.instructor = self.request.user
        form.instance.semester = Semester.objects.filter(is_active=True).first()
        response = super().form_valid(form)
        
        # Process all preference types using the same pattern
        self._process_preferences(
            'must_have_ta_', 
            MustHaveTAPreference, 
            'must_have_tas'
        )
        
        self._process_preferences(
            'preferred_ta_', 
            TAPreference, 
            'preferred_tas'
        )
        
        self._process_preferences(
            'preferred_grader_', 
            GraderPreference, 
            'preferred_graders'
        )
        
        self._process_preferences(
            'avoid_ta_', 
            AvoidTAPreference, 
            'tas_to_avoid'
        )
        
        self._process_preferences(
            'avoid_grader_', 
            AvoidGraderPreference, 
            'graders_to_avoid'
        )
        
        messages.success(self.request, 'TA request created successfully!')
        return response
    
    def _process_preferences(self, prefix, model_class, relation_name):
        """Process preferences from POST data and save them to the database"""
        preferences = {}
        # Get all preferences from POST data
        for key, value in self.request.POST.items():
            if key.startswith(prefix) and value:
                try:
                    order = int(key.split('_')[-1])
                    ta_id = int(value)
                    preferences[order] = ta_id
                except (ValueError, IndexError):
                    pass
        
        # Save preferences with order
        for order, ta_id in preferences.items():
            try:
                ta_profile = TAProfile.objects.get(id=ta_id)
                model_class.objects.create(
                    ta_request=self.object,
                    ta=ta_profile,
                    preference_order=order
                )
            except TAProfile.DoesNotExist:
                pass


class TARequestUpdateView(LoginRequiredMixin, InstructorRequiredMixin, UpdateView):
    """Update an existing TA request"""
    model = TARequest
    form_class = TARequestForm
    template_name = 'ta_request_cs_dept/request_form.html'
    success_url = reverse_lazy('ta_request_cs_dept:list')
    
    def dispatch(self, request, *args, **kwargs):
        # Check if request period is active for the semester this request belongs to
        ta_request = self.get_object()
        
        if not ta_request.semester.is_request_period_active():
            messages.error(request, f"TA request editing is closed. The request period ended on {ta_request.semester.request_end_date}.")
            return redirect('ta_request_cs_dept:list')
            
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        # Ensure instructors can only update their own requests
        return TARequest.objects.filter(instructor=self.request.user)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instructor'] = self.request.user
        kwargs['semester'] = self.object.semester
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tas'] = TAProfile.objects.filter(is_active=True)
        context['semester'] = self.object.semester
        
        # Add existing preferences to context
        preference_types = [
            ('must_have_tas', MustHaveTAPreference),
            ('preferred_tas', TAPreference),
            ('preferred_graders', GraderPreference),
            ('tas_to_avoid', AvoidTAPreference),
            ('graders_to_avoid', AvoidGraderPreference)
        ]
        
        for relation_name, model_class in preference_types:
            context[relation_name] = list(
                model_class.objects.filter(ta_request=self.object)
                .order_by('preference_order')
                .values('ta_id', 'preference_order')
            )
        
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Clear existing preferences
        MustHaveTAPreference.objects.filter(ta_request=self.object).delete()
        TAPreference.objects.filter(ta_request=self.object).delete()
        GraderPreference.objects.filter(ta_request=self.object).delete()
        AvoidTAPreference.objects.filter(ta_request=self.object).delete()
        AvoidGraderPreference.objects.filter(ta_request=self.object).delete()
        
        # Process all preference types
        self._process_preferences('must_have_ta_', MustHaveTAPreference)
        self._process_preferences('preferred_ta_', TAPreference)
        self._process_preferences('preferred_grader_', GraderPreference)
        self._process_preferences('avoid_ta_', AvoidTAPreference)
        self._process_preferences('avoid_grader_', AvoidGraderPreference)
        
        messages.success(self.request, 'TA request updated successfully!')
        return response
    
    def _process_preferences(self, prefix, model_class):
        """Process preferences from POST data and save them to the database"""
        preferences = {}
        for key, value in self.request.POST.items():
            if key.startswith(prefix) and value:
                try:
                    order = int(key.split('_')[-1])
                    ta_id = int(value)
                    preferences[order] = ta_id
                except (ValueError, IndexError):
                    pass
        
        # Save preferences with order
        for order, ta_id in preferences.items():
            try:
                ta_profile = TAProfile.objects.get(id=ta_id)
                model_class.objects.create(
                    ta_request=self.object,
                    ta=ta_profile,
                    preference_order=order
                )
            except TAProfile.DoesNotExist:
                pass


class SemesterListView(LoginRequiredMixin, CoordinatorRequiredMixin, ListView):
    """List all semesters for TA coordinators"""
    model = Semester
    template_name = 'ta_request_cs_dept/semester_list.html'
    context_object_name = 'semesters'
    ordering = ['-is_active', '-start_date']


class SemesterCreateView(LoginRequiredMixin, CoordinatorRequiredMixin, CreateView):
    """Create a new semester"""
    model = Semester
    form_class = SemesterForm
    template_name = 'ta_request_cs_dept/semester_form.html'
    success_url = reverse_lazy('ta_request_cs_dept:semester_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Semester created successfully!')
        return super().form_valid(form)


class SemesterUpdateView(LoginRequiredMixin, CoordinatorRequiredMixin, UpdateView):
    """Update semester details"""
    model = Semester
    form_class = SemesterForm
    template_name = 'ta_request_cs_dept/semester_form.html'
    success_url = reverse_lazy('ta_request_cs_dept:semester_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Semester updated successfully!')
        return super().form_valid(form)


class CoordinatorDashboardView(LoginRequiredMixin, CoordinatorRequiredMixin, ListView):
    """Dashboard for TA coordinators to manage all requests"""
    model = TARequest
    template_name = 'ta_request_cs_dept/coordinator_dashboard.html'
    context_object_name = 'ta_request_cs_dept'
    
    def get_queryset(self):
        # Filter by semester if provided
        semester_id = self.request.GET.get('semester')
        queryset = TARequest.objects.all()
        
        if semester_id:
            try:
                semester = Semester.objects.get(id=semester_id)
                queryset = queryset.filter(semester=semester)
            except Semester.DoesNotExist:
                pass
        else:
            # Default to active semester
            active_semester = Semester.objects.filter(is_active=True).first()
            if active_semester:
                queryset = queryset.filter(semester=active_semester)
        
        # Prefetch related data to improve performance
        return queryset.select_related('instructor', 'course', 'semester')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['semesters'] = Semester.objects.all()
        
        # Get current filter or default to active semester
        semester_id = self.request.GET.get('semester')
        if semester_id:
            context['current_semester_id'] = int(semester_id)
        else:
            active_semester = Semester.objects.filter(is_active=True).first()
            if active_semester:
                context['current_semester_id'] = active_semester.id
        
        return context


class TARequestDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Detailed view of a TA request for coordinators and request owners"""
    model = TARequest
    template_name = 'ta_request_cs_dept/request_detail.html'
    context_object_name = 'ta_request'
    
    def test_func(self):
        obj = self.get_object()
        # Allow access if user is request owner or coordinator
        return (self.request.user == obj.instructor) or \
               (self.request.user.is_instructor() and 
                getattr(self.request.user.instructor_profile, 'is_ta_coordinator', False))


def export_requests_csv(request):
    """Export all TA requests for a semester as CSV"""
    if not request.user.is_authenticated:
        return redirect('accounts:login')
        
    if not (request.user.is_instructor() and 
           getattr(request.user.instructor_profile, 'is_ta_coordinator', False)):
        messages.error(request, "You don't have permission to export TA requests.")
        return redirect('home')
    
    # Get semester from query param or use active semester
    semester_id = request.GET.get('semester')
    if semester_id:
        try:
            semester = Semester.objects.get(id=semester_id)
        except Semester.DoesNotExist:
            semester = Semester.objects.filter(is_active=True).first()
    else:
        semester = Semester.objects.filter(is_active=True).first()
    
    if not semester:
        messages.error(request, "No active semester found.")
        return redirect('ta_request_cs_dept:coordinator_dashboard')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ta_request_cs_dept_{semester.name.replace(" ", "_")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Instructor', 'Course Code', 'Course Name', 'Min TA Loads', 'Max TA Loads', 
        'Graders', 'Must-Have TAs', 'Preferred TAs', 'Preferred Graders',
        'TAs to Avoid', 'Graders to Avoid', 'Must-Have Justification', 'General Justification'
    ])
    
    # Query all requests for this semester with prefetched data
    ta_request_cs_dept = TARequest.objects.filter(semester=semester).select_related(
        'instructor', 'course'
    ).prefetch_related(
        'must_have_tas__user', 'preferred_tas__user', 
        'preferred_graders__user', 'tas_to_avoid__user', 'graders_to_avoid__user'
    )
    
    for req in ta_request_cs_dept:
        # Get lists of TAs in preference order
        must_have_tas = _get_tas_in_order(req, MustHaveTAPreference)
        preferred_tas = _get_tas_in_order(req, TAPreference)
        preferred_graders = _get_tas_in_order(req, GraderPreference)
        avoid_tas = _get_tas_in_order(req, AvoidTAPreference)
        avoid_graders = _get_tas_in_order(req, AvoidGraderPreference)
        
        writer.writerow([
            req.instructor.get_full_name(),
            req.course.code,
            req.course.title,
            req.min_ta_loads,
            req.max_ta_loads,
            req.graders_requested,
            '; '.join(must_have_tas),
            '; '.join(preferred_tas),
            '; '.join(preferred_graders),
            '; '.join(avoid_tas),
            '; '.join(avoid_graders),
            req.must_have_justification,
            req.general_justification
        ])
    
    return response


def export_requests_excel(request):
    """Export all TA requests for a semester as Excel"""
    if not request.user.is_authenticated:
        return redirect('accounts:login')
        
    if not (request.user.is_instructor() and 
           getattr(request.user.instructor_profile, 'is_ta_coordinator', False)):
        messages.error(request, "You don't have permission to export TA requests.")
        return redirect('home')
    
    # Get semester from query param or use active semester
    semester_id = request.GET.get('semester')
    if semester_id:
        try:
            semester = Semester.objects.get(id=semester_id)
        except Semester.DoesNotExist:
            semester = Semester.objects.filter(is_active=True).first()
    else:
        semester = Semester.objects.filter(is_active=True).first()
    
    if not semester:
        messages.error(request, "No active semester found.")
        return redirect('ta_request_cs_dept:coordinator_dashboard')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TA Requests"
    
    # Write headers
    headers = [
        'Instructor', 'Course Code', 'Course Name', 'Min TA Loads', 'Max TA Loads', 
        'Graders', 'Must-Have TAs', 'Preferred TAs', 'Preferred Graders',
        'TAs to Avoid', 'Graders to Avoid', 'Must-Have Justification', 'General Justification'
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Query all requests
    ta_request_cs_dept = TARequest.objects.filter(semester=semester).select_related(
        'instructor', 'course'
    ).prefetch_related(
        'must_have_tas__user', 'preferred_tas__user', 
        'preferred_graders__user', 'tas_to_avoid__user', 'graders_to_avoid__user'
    )
    
    # Write data
    for row_num, req in enumerate(ta_request_cs_dept, 2):
        # Get lists of TAs in preference order
        must_have_tas = _get_tas_in_order(req, MustHaveTAPreference)
        preferred_tas = _get_tas_in_order(req, TAPreference)
        preferred_graders = _get_tas_in_order(req, GraderPreference)
        avoid_tas = _get_tas_in_order(req, AvoidTAPreference)
        avoid_graders = _get_tas_in_order(req, AvoidGraderPreference)
        
        row_data = [
            req.instructor.get_full_name(),
            req.course.code,
            req.course.title,
            req.min_ta_loads,
            req.max_ta_loads,
            req.graders_requested,
            '; '.join(must_have_tas),
            '; '.join(preferred_tas),
            '; '.join(preferred_graders),
            '; '.join(avoid_tas),
            '; '.join(avoid_graders),
            req.must_have_justification,
            req.general_justification
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Create response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ta_request_cs_dept_{semester.name.replace(" ", "_")}.xlsx"'
    
    return response


def _get_tas_in_order(ta_request, preference_model):
    """
    Helper function to get TAs in preference order as formatted strings
    """
    tas = preference_model.objects.filter(
        ta_request=ta_request
    ).order_by('preference_order').select_related('ta__user')
    
    return [
        f"{i+1}. {pref.ta.user.get_full_name() or pref.ta.user.username}"
        for i, pref in enumerate(tas)
    ]


def get_tas_json(request):
    """API endpoint to get TAs as JSON for AJAX requests"""
    tas = TAProfile.objects.filter(is_active=True)
    
    # Optional search query
    query = request.GET.get('q', '')
    if query:
        tas = tas.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__username__icontains=query)
        )
    
    tas_list = [{
        'id': ta.id,
        'name': f"{ta.user.get_full_name() or ta.user.username} ({ta.get_ta_type_display()})",
        'type': ta.get_ta_type_display()
    } for ta in tas]
    
    return JsonResponse(tas_list, safe=False)